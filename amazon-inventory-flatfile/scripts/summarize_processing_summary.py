#!/usr/bin/env python3
import argparse
import os
from pathlib import Path

from openpyxl import load_workbook


def parse_txt_summary(path):
    processed = successful = 0
    rows = []
    with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
        lines = [line.rstrip("\n") for line in f]

    in_table = False
    for line in lines:
        if line.startswith("\tNumber of records processed"):
            processed = int(line.split("\t")[-1] or 0)
        elif line.startswith("\tNumber of records successful"):
            successful = int(line.split("\t")[-1] or 0)
        elif line.startswith("original-record-number"):
            in_table = True
        elif in_table and line:
            parts = line.split("\t")
            if len(parts) >= 5:
                rows.append(
                    {
                        "error_code": parts[2],
                        "error_type": parts[3],
                        "message": parts[4],
                        "sku": parts[1],
                    }
                )
    return {
        "processed": processed,
        "successful": successful,
        "success_with_other_errors": 0,
        "failed": max(processed - successful, 0),
        "warning_only": 0,
        "error_rows": rows,
        "sku_rows": [row for row in rows if row.get("sku")],
    }


def parse_xlsm_summary(path):
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    result_ws = wb["フィード処理結果"]
    template_ws = wb["テンプレート"]

    processed = int(result_ws.cell(4, 5).value or 0)
    successful = int(result_ws.cell(5, 5).value or 0)
    success_with_other_errors = int(result_ws.cell(6, 5).value or 0)
    failed = int(result_ws.cell(7, 5).value or 0)
    warning_only = int(result_ws.cell(8, 5).value or 0)

    error_rows = []
    row_num = 12
    while True:
        code = result_ws.cell(row_num, 3).value
        if row_num > 12 and not code:
            break
        if code and str(code) != "None" and str(code) != "エラーコード":
            error_rows.append(
                {
                    "error_code": str(code),
                    "error_type": str(result_ws.cell(row_num, 4).value or ""),
                    "message": str(result_ws.cell(row_num, 5).value or ""),
                    "field": str(result_ws.cell(row_num, 6).value or ""),
                    "count": str(result_ws.cell(row_num, 7).value or ""),
                }
            )
        row_num += 1
        if row_num > 200:
            break

    sku_rows = []
    row_num = 21
    while True:
        code = result_ws.cell(row_num, 3).value
        sku = result_ws.cell(row_num, 7).value
        if row_num > 21 and not code and not sku:
            break
        if code and sku:
            sku_rows.append(
                {
                    "error_code": str(code),
                    "error_type": str(result_ws.cell(row_num, 4).value or ""),
                    "message": str(result_ws.cell(row_num, 5).value or ""),
                    "field": str(result_ws.cell(row_num, 6).value or ""),
                    "sku": str(sku),
                }
            )
        row_num += 1
        if row_num > 300:
            break

    statuses = []
    row_num = 7
    while True:
        status = template_ws.cell(row_num, 1).value
        sku = template_ws.cell(row_num, 4).value
        if row_num > 7 and not status and not sku:
            break
        if sku:
            statuses.append(
                {
                    "status": str(status or ""),
                    "sku": str(sku),
                    "fulfillment": str(template_ws.cell(row_num, 5).value or ""),
                    "qty": str(template_ws.cell(row_num, 6).value or ""),
                }
            )
        row_num += 1
        if row_num > 10000:
            break

    return {
        "processed": processed,
        "successful": successful,
        "success_with_other_errors": success_with_other_errors,
        "failed": failed,
        "warning_only": warning_only,
        "error_rows": error_rows,
        "sku_rows": sku_rows,
        "statuses": statuses,
    }


def classify(message):
    lowered = message.lower()
    if "最低価格" in message or "minimum" in lowered:
        return "Price issue"
    if "画像" in message or "media" in lowered or "image" in lowered:
        return "Image issue"
    if "manufacturer" in lowered or "カタログ" in message:
        return "Catalog conflict"
    return "Other issue"


def render_markdown(summary_path, parsed):
    lines = []
    lines.append("# Amazon Inventory Upload Execution Report")
    lines.append("")
    lines.append(f"Source summary: `{summary_path}`")
    lines.append("")
    lines.append("## Inventory Update Result")
    lines.append("")
    lines.append(f"- Total SKUs processed: `{parsed['processed']}`")
    lines.append(f"- Successful SKUs: `{parsed['successful']}`")
    lines.append(
        f"- Successful with other errors: `{parsed.get('success_with_other_errors', 0)}`"
    )
    lines.append(f"- Failed SKUs: `{parsed['failed']}`")
    lines.append(f"- Warning-only SKUs: `{parsed['warning_only']}`")
    lines.append("")

    if parsed["failed"] == 0:
        lines.append("Conclusion:")
        lines.append("")
        lines.append("- The inventory upload completed without failed rows.")
        lines.append("- Any additional issues are listing-level follow-up items.")
        lines.append("")

    if parsed["sku_rows"]:
        lines.append("## Other Errors Found")
        lines.append("")
        grouped = {}
        for row in parsed["sku_rows"]:
            key = (row["error_code"], classify(row["message"]), row["message"])
            grouped.setdefault(key, []).append(row["sku"])
        for (error_code, label, message), skus in grouped.items():
            lines.append(f"### {label}")
            lines.append("")
            lines.append(f"- Error code: `{error_code}`")
            lines.append(f"- Message: {message}")
            lines.append("- Affected SKUs:")
            for sku in skus:
                lines.append(f"  - `{sku}`")
            lines.append("")
            lines.append("Recommendation:")
            if label == "Price issue":
                lines.append("")
                lines.append("- Review minimum/maximum price guardrails in Seller Central.")
            elif label == "Image issue":
                lines.append("")
                lines.append("- Fix the media source or file format outside the inventory feed.")
            elif label == "Catalog conflict":
                lines.append("")
                lines.append("- Confirm the ASIN mapping and resolve the catalog conflict with Amazon if needed.")
            else:
                lines.append("")
                lines.append("- Review the listing manually before the next upload.")
            lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--summary-path", required=True)
    parser.add_argument("--output-path")
    args = parser.parse_args()

    summary_path = Path(args.summary_path)
    suffix = summary_path.suffix.lower()
    if suffix == ".xlsm":
        parsed = parse_xlsm_summary(summary_path)
    else:
        parsed = parse_txt_summary(summary_path)

    output_path = (
        Path(args.output_path)
        if args.output_path
        else summary_path.with_name(summary_path.stem + "_report.md")
    )
    output_path.write_text(render_markdown(summary_path, parsed), encoding="utf-8")
    print(output_path)


if __name__ == "__main__":
    main()
