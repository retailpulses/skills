#!/usr/bin/env python3
"""Build Amazon inventory flatfile from Supabase amazon_listings table.

Replaces Baserow API with Supabase PostgREST.
Reads from amazon_listings (domain: product_catalog, owner: retailpulses/RPagentOS).
"""
import argparse
import csv
import json
import os
import urllib.request

from openpyxl import load_workbook

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")


def fetch_rows(table_name, limit=200):
    """Fetch all rows from a Supabase table via PostgREST."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise SystemExit("Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")
    rows = []
    offset = 0
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Accept": "application/json",
    }
    while True:
        url = f"{SUPABASE_URL.rstrip('/')}/rest/v1/{table_name}?select=*&limit={limit}&offset={offset}"
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=60) as resp:
            batch = json.loads(resp.read().decode())
        if not batch:
            break
        rows.extend(batch if isinstance(batch, list) else [batch])
        if len(batch) < limit:
            break
        offset += limit
    return rows


def max_header_columns(ws):
    last = 0
    for r in range(1, 7):
        for c in range(1, ws.max_column + 1):
            value = ws.cell(r, c).value
            if value not in (None, ""):
                last = max(last, c)
    return last


def normalized_order(row):
    try:
        return float(row.get("order") or row["id"])
    except Exception:
        return float(row["id"])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--supabase-url", default=SUPABASE_URL)
    parser.add_argument("--supabase-key", default=SUPABASE_KEY)
    parser.add_argument("--template-path", required=True)
    parser.add_argument("--output-path", required=True)
    parser.add_argument("--sku-field", default="seller_sku")
    parser.add_argument(
        "--fulfillment-field", default="fulfillment_channel"
    )
    parser.add_argument("--qty-field", default="quantity")
    args = parser.parse_args()

    global SUPABASE_URL, SUPABASE_KEY
    SUPABASE_URL = args.supabase_url or SUPABASE_URL
    SUPABASE_KEY = args.supabase_key or SUPABASE_KEY

    wb = load_workbook(args.template_path, read_only=True, data_only=False, keep_vba=True)
    ws = wb["テンプレート"]
    col_count = max_header_columns(ws)
    header_rows = []
    for r in range(1, 7):
        header_rows.append(
            [
                ws.cell(r, c).value if ws.cell(r, c).value is not None else ""
                for c in range(1, col_count + 1)
            ]
        )

    amazon_rows = fetch_rows("amazon_listings")
    amazon_rows.sort(key=normalized_order)

    written = 0
    samples = []
    with open(args.output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter="\t", lineterminator="\n")
        writer.writerows(header_rows)
        for source in amazon_rows:
            sku = str(source.get(args.sku_field) or "").strip()
            if not sku:
                continue
            row = [""] * col_count
            row[0] = sku
            row[1] = str(source.get(args.fulfillment_field) or "").strip()
            row[2] = str(source.get(args.qty_field) or "").strip()
            writer.writerow(row)
            written += 1
            if len(samples) < 10:
                samples.append(row[:3])

    print(
        json.dumps(
            {
                "output_path": args.output_path,
                "data_rows_written": written,
                "template_columns": col_count,
                "sample_rows": samples,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
